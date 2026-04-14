#!/usr/bin/env python3
"""Web UI for imgdl — local Flask server with image download, batch upload, and templates."""

import io
import json
import os
import tempfile
import threading
from pathlib import Path

from flask import Flask, Response, jsonify, request, send_file, send_from_directory

from imgdl import (
    DownloadOpts,
    PRODUCT_TYPES,
    QueryItem,
    WINE_TYPES,
    load_config,
    load_queries,
    parse_pasted_text,
    parse_size,
    run_batch,
)


def _build_opts(data: dict, is_form: bool = False) -> DownloadOpts:
    """Build DownloadOpts from request data (JSON or form), with auto-defaults for wine types."""
    get = (lambda k, d=None: data.get(k, d)) if not is_form else (lambda k, d=None: data.get(k, d))
    try:
        size = parse_size(get("size", "300x300"))
    except Exception:
        size = (300, 300)

    img_type = get("type") or None
    min_src = 70 if img_type in (WINE_TYPES | PRODUCT_TYPES) else 0

    return DownloadOpts(
        size=size,
        type=img_type,
        background=get("background") or None,
        format=get("format", "png"),
        count=max(1, min(int(get("count", 1)), 5)),
        padding=int(get("padding", 0)),
        output=str(DOWNLOAD_DIR),
        min_source_pct=min_src,
    )

app = Flask(__name__, static_folder="static")

DOWNLOAD_DIR = Path("./downloads")
DOWNLOAD_DIR.mkdir(exist_ok=True)

# Load API keys once at startup
_cfg = load_config()
_sources_cfg = _cfg.get("sources", {})
GOOGLE_KEY = os.environ.get("GOOGLE_CSE_API_KEY") or _sources_cfg.get("google_cse_key")
GOOGLE_ID = os.environ.get("GOOGLE_CSE_ID") or _sources_cfg.get("google_cse_id")
BRAVE_KEY = os.environ.get("BRAVE_API_KEY") or _sources_cfg.get("brave_api_key")


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/downloads/<path:filename>")
def serve_download(filename):
    return send_from_directory(DOWNLOAD_DIR, filename)


@app.route("/api/download", methods=["POST"])
def api_download():
    """Single query or URL download via JSON body. Returns SSE stream."""
    data = request.get_json(force=True)
    query_str = data.get("query", "").strip()
    if not query_str:
        return jsonify({"error": "query is required"}), 400

    opts = _build_opts(data)
    items = load_queries(query_str)
    return _stream_batch(items, opts)


@app.route("/api/upload", methods=["POST"])
def api_upload():
    """Batch download from uploaded file (CSV/XLSX/TXT). Returns SSE stream."""
    if "file" not in request.files:
        return jsonify({"error": "no file uploaded"}), 400

    uploaded = request.files["file"]
    suffix = Path(uploaded.filename).suffix.lower()
    if suffix not in (".txt", ".csv", ".xlsx", ".xls"):
        return jsonify({"error": f"unsupported file type: {suffix}"}), 400

    # Save to temp file so load_queries can read it
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    uploaded.save(tmp.name)
    tmp.close()

    try:
        items = load_queries(tmp.name)
    finally:
        os.unlink(tmp.name)

    if not items:
        return jsonify({"error": "no queries found in file"}), 400

    opts = _build_opts(dict(request.form))
    return _stream_batch(items, opts)


@app.route("/api/paste", methods=["POST"])
def api_paste():
    """Batch download from pasted text. Cleans numbered lists, URLs, brackets."""
    data = request.get_json(force=True)
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "no text provided"}), 400

    items = parse_pasted_text(text)
    if not items:
        return jsonify({"error": "could not parse any queries from pasted text"}), 400

    opts = _build_opts(data)
    return _stream_batch(items, opts)


def _stream_batch(items, opts):
    """Return an SSE Response that streams progress for a batch."""
    def generate():
        final_files = []
        for progress in run_batch(items, opts, GOOGLE_KEY, GOOGLE_ID, BRAVE_KEY):
            final_files.extend(progress["stats"].get("files", []))
            yield f"data: {json.dumps(progress)}\n\n"
        # Final summary event
        yield f"data: {json.dumps({'done': True, 'files': final_files, 'cumulative': progress['cumulative']})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# Template downloads
# ---------------------------------------------------------------------------

TEMPLATE_HEADERS = ["query", "url", "size", "type", "background", "format", "filename"]
TEMPLATE_ROWS = [
    ["Anthropic logo", "", "300x300", "logo", "transparent", "png", ""],
    ["Tesla Model 3", "", "500x500", "product photo", "white", "jpg", ""],
    ["", "https://example.com/image.png", "200x200", "", "", "png", "custom_name"],
]


@app.route("/api/template/csv")
def template_csv():
    buf = io.StringIO()
    import csv
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerows(TEMPLATE_ROWS)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=imgdl_template.csv"},
    )


@app.route("/api/template/xlsx")
def template_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Queries"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")

    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(TEMPLATE_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="imgdl_template.xlsx")


# ---------------------------------------------------------------------------
# List downloaded files
# ---------------------------------------------------------------------------

@app.route("/api/files")
def list_files():
    """List all downloaded image files."""
    files = sorted(
        [f.name for f in DOWNLOAD_DIR.iterdir()
         if f.is_file() and f.suffix.lower() in (".png", ".jpg", ".jpeg", ".webp")],
    )
    return jsonify({"files": files})


if __name__ == "__main__":
    print("Starting imgdl web UI → http://localhost:5001")
    app.run(host="127.0.0.1", port=5001, debug=True)
